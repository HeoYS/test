
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "=SUM(5, 10, 15)"
ws["A2"] = "=AVERAGE(5, 2, 3)"

ws["A3"] = 20
ws["A4"] = 30
ws["A5"] = "=SUM(A3:A4)"

wb.save("formula.xlsx")