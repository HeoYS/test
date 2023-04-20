from openpyxl import load_workbook

wb = load_workbook('test2.xlsx')
ws = wb.active

#cell 데이터 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(x, y).value, end=' ')
    print()

#cell 갯수 모를때
for x in range(1, ws.max_row):
    for y in range(1, ws.max_column):
        print(ws.cell(x, y).value, end=' ')
    print()

