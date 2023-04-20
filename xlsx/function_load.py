#수식 적용된 셀의 데이터 가져오기

from openpyxl import load_workbook
# (function) load_workbook: (filename, read_only=False, keep_vba=KEEP_VBA,data_only=False, keep_links=True) -> Workbook
wb = load_workbook("formula.xlsx", data_only=True) #data_only : 수식 불러오기 말고 값만 불러오기, None 뜨면 엑셀파일 들어가서 저장누르
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)

# wb.save("formula.xlsx")