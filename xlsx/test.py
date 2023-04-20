
from openpyxl import Workbook


wb = Workbook() #새 엑셀파일 생성
ws = wb.create_sheet() #새 시트 생성
ws.title = 'MySheet' #시트 이름 변경
ws.sheet_properties.tabColor = 'ff66ff' #시트 탭 색 변경

ws1 = wb.create_sheet('YourSheet') #YourSheet이름으로 시트 생성
ws2 = wb.create_sheet('NewSheet', 2) #두번째에 시트 생성

#시트 접근
print(wb['NewSheet'].title)
new_ws = wb['NewSheet']
print(new_ws.title)

print(wb.sheetnames) #모든 시트 확인, 리스트로 반환

#시트 복사
new_ws['A1'] = 'Test' #A1에 데이터 넣음
target = wb.copy_worksheet(new_ws) #복사된 시트가 우측 마짖막에 생성됨(데이터 포함)
target.title = 'Copied_Sheet'

wb.save('test2.xlsx')