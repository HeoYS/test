
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection

wb = load_workbook("test.xlsx")

ws = wb["New Sheet"]

a1 = ws.cell(row=1, column=1, value="파이썬은")
b1 = ws.cell(row=1, column=2, value="역시")
c1 = ws.cell(row=1, column=3, value="족같다")

a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="0000FF", name="Arial", strike=True) #취소선 적용
c1.font = Font(color="00FF00", size=15, underline="single") #밑줄 적용

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
# protection = Protection(locked=True, hidden=False) 엑셀파일 보호

wb.save("test.xlsx") #엑셀파일 저장