#저장된 엑셀 불러오기

from openpyxl import load_workbook

wb = load_workbook("test.xlsx")

ws = wb["First Sheet"]
#값 저장

ws["A1"] = 5
ws.cell(row=1, column=2, value=70)
# ws.cell(1,2,value=70)
ws["A1"] = "" #셀삭제
ws.delete_rows(1) #1번행 1개 삭제
ws.delete_rows(1,3) #1번행부터 총 3개의 행 삭제
ws.delete_cols(3) #3번열 1개 삭제

ws.insert_rows(1) #1번행에 빈행 삽입

ws["D1"] = 3
ws.move_range("D1", cols=2) #D1에 있는걸 오른쪽으로 2열 이동
ws.move_range("D4:F10", rows=1, cols=2) #D4~F10 영역의 데이터를 위로 1행 오른쪽으로 2열 이동

ws["A1"] = 3
#A1에 있는 셀 데이터 가져오기
ws['A1'].value
ws.cell(row=1, column=1).value

ws['A4'] = 10
ws.cell(1, 2, value=15)

#해당 위치의 데이터 확인
a = ws['A4']
print(a)
b = ws['A4'].value
print(b)
c = ws.cell(row=1, column=2)
print(c)
d = ws.cell(row=1, column=2).value
print(d)

#A1:B4 범위의 셀 정보 가져오기
cell_range = ws['A1:B4']
for i in cell_range:
    for j in i:
        print(j.value)

#행 단위로 가져오기
for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=2):
    for cell in row:
        print(cell)
#열 단위로 가져오기
for col in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=2):
    for cell in col:
        print(cell)
#열 하나 가져오기
col_C = ws['C']
for cell in col_C:
    print(cell.value)
#열 여러개 가져오기
col_range = ws['B:C']
for cols in col_range:
    for cell in cols:
        print(cell.value)
#행 하나 가져오기 row10 = ws[10]
#행 여러개 가져오기 row_range = ws[5:10]

#모든 행이나 열을 참조할때
ws = wb.active
ws['C9'] = 'hello world'
tuple(ws.rows)
tuple(ws.columns)



wb.save("test.xlsx") #엑셀파일 저장