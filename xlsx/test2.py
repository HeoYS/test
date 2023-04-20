import openpyxl
from openpyxl import Workbook
from random import *

wb = openpyxl.load_workbook(filename='test2.xlsx')
ws = wb.active #첫번째 시트 활성

ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3
ws['A4'] = 4

print(ws['A1']) # 셀 객체 정보만 출력
print(ws['A1'].value) # 입력된 값 출력
print(ws['A10'].value) # 없으면 None 출력

print(ws.cell(row=1, column=1).value)
print(ws.cell(1, 1).value) #위에랑 이거랑 같음

ws.cell(1, 3).value = 10
ws.cell(2, 3, value=20)
c = ws.cell(3, 3, value=30)
print(c.value)

for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(x, y).value = randint(0, 100) #0~100사이 숫자

wb.save('test2.xlsx')



